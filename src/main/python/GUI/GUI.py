import os, sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from matplotlib.pyplot import *
import openpyxl
from PIL import Image
from PIL.ImageQt import ImageQt
from matplotlib.backends.backend_agg import FigureCanvasAgg
import io
import numpy as np
import Controller


class Window(QMainWindow):
    def __init__(self):
        self.width = 1280
        self.height = 720
        super().__init__()

        self.currentShopTab = None
        self.currentCardTab = None

        self.shopsTabName = "Магазины"
        self.cardsTabName = "Карты"
        self.backButtonName = "Назад"
        self.commonStatsName = "Общая статистика магазинов по городам"
        self.shopsStatsName = "Статистика по магазинам"
        self.cardsStatsName = "Статистика по картам"
        self.chooseTimrStr = "Выберете преиод времени"
        self.yearsStr = "Года"
        self.yearStr = "Год"
        self.monthsStr = "Месяцы"
        self.monthStr = "Месяц"
        self.weeksStr = "Недели"
        self.weekStr = "Неделя"
        self.showTableStr = "Показать таблицу"
        self.cardStatsLabels = ["Количество чеков с картами",
                                "Отношение количества чеков с картами к общему числу чеков",
                                "Общая сумма продаж в чеках с картами",
                                "Отношение суммы продаж в чеках с картами к общей сумме"]
        # labels = ["Conversion", "Units per transaction",
        #           "Average check", "Sales per area", "Count of checks", "Returned units", "Count of visitors",
        #           "Proceeds without tax", "Proceeds with tax", "Count of sold units"]
        self.shopStatsLabels = ["Конверсия", "Среднее число товаров в чеке", "Средний чек", "Продажи с кв. метра",
                                "Количество чеков", "Количество возвратов", "Число посетителей", "Доход без НДС",
                                "Доход с НДС", "Число проданных товаров"]
        self.itemPairsLabels = ["Товар №1", "Товар №2", "Частота"]
        self.itemLabels = ["Товар", "Частота"]
        self.frequencyStr = "Частота"
        self.countStr = "Количество"
        self.manLabel = "Мужчины"
        self.womanLabel = "Женщины"

        self.chooseIndexStr = "Выберете показатель:"

        self.chooseDirectoryStr = "Выберете папку"
        self.chooseNameStr = "Выберете имя"
        self.enterPngNameStr = "Введите имя .png файла"
        self.enterExcelNameStr = "Введите имя .xlsx файла"
        self.fileNameNotChosenMsg = "Имя файла не выбрано"

        self.toExcelButtonName = "Сохранить в .xlsx файл"
        self.viewDiagramButtonName = "Посмотреть диаграмму"
        self.toPngButtonName = "Сохранить в .png файл"

        self.controller = Controller.Controller()

        self.initUi()

    def initUi(self):
        self.setFixedSize(self.width, self.height)
        self.center()
        self.setWindowTitle('RetailDB')

        self.shopDefaultWindow(0)

        self.show()

    def center(self):
        frameGm = self.frameGeometry()
        centerPoint = QDesktopWidget().availableGeometry().center()
        frameGm.moveCenter(centerPoint)
        self.move(frameGm.topLeft())

    def shopDefaultWindow(self, activeTab):
        tabBar = QTabWidget(self)
        tabBar.resize(self.width, self.height)
        tab1 = QWidget()
        tab2 = QWidget()
        tabBar.addTab(tab1, self.shopsTabName)
        tabBar.addTab(tab2, self.cardsTabName)

        if activeTab == 0:
            self.currentShopTab = [tab1]
        elif activeTab == 1:
            self.currentCardTab = [tab2]
        else:
            raise RuntimeError("active tab > 1")

        tabBar.setCurrentIndex(activeTab)

        commonStats1 = QPushButton(self.commonStatsName)
        shopStats1 = QPushButton(self.shopsStatsName)

        shopStats1.pressed.connect(lambda: self.shopChooseShopWindow(False))
        commonStats1.pressed.connect(lambda: self.shopChooseShopWindow(True))

        vbox1 = QVBoxLayout(tab1)
        vbox1.addWidget(shopStats1)
        vbox1.addWidget(commonStats1)
        vbox1.addStretch()

        shopStats2 = QPushButton(self.cardsStatsName)

        shopStats2.pressed.connect(lambda: self.shopChooseCardWindow())

        vbox2 = QVBoxLayout(tab2)
        vbox2.addWidget(shopStats2)
        vbox2.addStretch()

        self.currentShopTab = tab1
        self.currentCardTab = tab2

        tabBar.show()

    def shopChooseCardWindow(self, start=0):
        tabBar = QTabWidget(self)

        tabBar.resize(self.width, self.height)
        tab1 = self.currentShopTab
        tab2 = QWidget()
        tabBar.addTab(tab1, self.shopsTabName)
        tabBar.addTab(tab2, self.cardsTabName)

        tabBar.setCurrentIndex(1)

        self.currentCardTab = tab2

        back = QPushButton(self.backButtonName)
        back.setFixedWidth(50)

        back.pressed.connect(lambda: self.shopDefaultWindow(1))

        itemList = self.controller.getShopNames()

        tree = QTreeWidget()
        tree.setHeaderLabel(self.chooseTimrStr)

        parent = QTreeWidgetItem(tree)

        parent.setText(0, self.yearsStr)
        parent.setFlags(parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
        #years = [2015, 2016, 2017, 2018]
        #months = [1, 2, 3, 4, 5]
        #weeks = [1, 2, 3]
        #years, months, weeks = self.controller.getCardTimes()

        years, months, weeks = self.controller.getCardTimes(itemList[start])

        shopList = QComboBox()
        shopList.addItems(itemList)
        shopList.setCurrentIndex(start)
        shopList.currentIndexChanged.connect(lambda: self.shopChooseCardWindow(shopList.currentIndex()))

        for year in years:
            child = QTreeWidgetItem(parent)
            child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
            child.setText(0, str(year))
            child.setCheckState(0, Qt.Unchecked)

        parent = QTreeWidgetItem(tree)
        parent.setText(0, self.monthsStr)
        parent.setFlags(parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
        for month in months:
            child = QTreeWidgetItem(parent)
            child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
            child.setText(0, str(month))
            child.setCheckState(0, Qt.Unchecked)

        parent = QTreeWidgetItem(tree)
        parent.setText(0, self.weeksStr)
        parent.setFlags(parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
        for week in weeks:
            child = QTreeWidgetItem(parent)
            child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
            child.setText(0, str(week))
            child.setCheckState(0, Qt.Unchecked)

        showTable = QPushButton(self.showTableStr)
        showTable.pressed.connect(lambda: self.showTableCardWindow(
            itemList[shopList.currentIndex()],
            [item for item in tree.topLevelItem(0).takeChildren() if item.checkState(0) > 0],
            [item for item in tree.topLevelItem(1).takeChildren() if item.checkState(0) > 0],
            [item for item in tree.topLevelItem(2).takeChildren() if item.checkState(0) > 0]))

        showTable.setDisabled(True)

        def checkDates():
            def checkOneYear():
                count = 0
                for i in range(tree.topLevelItem(0).childCount()):
                    if tree.topLevelItem(0).child(i).checkState(0) > 0:
                        count += 1

                if count > 1:
                    return False
                elif count == 1:
                    return True
                else:
                    return False

            res = True
            res = res and tree.topLevelItem(0).checkState(0) > 0
            res = res and (
                (((tree.topLevelItem(1).checkState(0) > 0) ^ (tree.topLevelItem(2).checkState(0) > 0)) and (checkOneYear()))
                or ((tree.topLevelItem(1).checkState(0) == 0) and (tree.topLevelItem(2).checkState(0) == 0)))

            return res

        tree.itemClicked.connect(lambda: showTable.setDisabled(False) if checkDates() else showTable.setDisabled(True))

        vbox = QVBoxLayout(tab2)
        vbox.addWidget(tree)
        vbox.addWidget(shopList)
        vbox.addWidget(showTable)
        vbox.addStretch(1)
        vbox.addWidget(back)

        tabBar.show()

    def shopChooseShopWindow(self, isCommon, start=0):
        tabBar = QTabWidget(self)
        tabBar.resize(self.width, self.height)
        tab1 = QWidget()
        tab2 = self.currentCardTab
        tabBar.addTab(tab1, self.shopsTabName)
        tabBar.addTab(tab2, self.cardsTabName)

        self.currentShopTab = tab1

        back = QPushButton(self.backButtonName)
        back.setFixedWidth(50)
        back.pressed.connect(lambda: self.shopDefaultWindow(0))

        tree = QTreeWidget()
        tree.setHeaderLabel(self.chooseTimrStr)

        if isCommon:
            #itemList = ["The whole company", "Moscow", "St. Petersburg", "Riga"]
            itemList = self.controller.getCities()
        else:
            #itemList = ["Moscow 1","Moscow 2", "Moscow 3"]
            itemList = self.controller.getShopNames()

        parent = QTreeWidgetItem(tree)
        parent.setText(0, self.yearsStr)
        parent.setFlags(parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
        #years = [2015, 2016, 2017, 2018]
        #months = [1, 2, 3, 4, 5]
        #weeks = [1, 2, 3]

        if isCommon:
            years, months, weeks = self.controller.getCommonShopTimes(itemList[start])
        else:
            years, months, weeks = self.controller.getShopTimes(itemList[start])

        shopList = QComboBox()
        shopList.addItems(itemList)
        shopList.setCurrentIndex(start)
        shopList.currentIndexChanged.connect(lambda: self.shopChooseShopWindow(isCommon, shopList.currentIndex()))

        for year in years:
            child = QTreeWidgetItem(parent)
            child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
            child.setText(0, str(year))
            child.setCheckState(0, Qt.Unchecked)

        parent = QTreeWidgetItem(tree)
        parent.setText(0, self.monthsStr)
        parent.setFlags(parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
        for month in months:
            child = QTreeWidgetItem(parent)
            child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
            child.setText(0, str(month))
            child.setCheckState(0, Qt.Unchecked)

        parent = QTreeWidgetItem(tree)
        parent.setText(0, self.weeksStr)
        parent.setFlags(parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
        for week in weeks:
            child = QTreeWidgetItem(parent)
            child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
            child.setText(0, str(week))
            child.setCheckState(0, Qt.Unchecked)

        showTable = QPushButton(self.showTableStr)
        showTable.pressed.connect(lambda: self.showTableShopWindow(
            itemList[shopList.currentIndex()],
            [item for item in tree.topLevelItem(0).takeChildren() if item.checkState(0) > 0],
            [item for item in tree.topLevelItem(1).takeChildren() if item.checkState(0) > 0],
            [item for item in tree.topLevelItem(2).takeChildren() if item.checkState(0) > 0],
            isCommon))

        showTable.setDisabled(True)

        def checkDates():
            def checkOneYear():
                count = 0
                for i in range(tree.topLevelItem(0).childCount()):
                    if tree.topLevelItem(0).child(i).checkState(0) > 0:
                        count += 1

                if count > 1:
                    return False
                elif count == 1:
                    return True
                else:
                    return False

            res = True
            res = res and tree.topLevelItem(0).checkState(0) > 0
            res = res and (
                (((tree.topLevelItem(1).checkState(0) > 0) ^ (tree.topLevelItem(2).checkState(0) > 0)) and (checkOneYear()))
                or ((tree.topLevelItem(1).checkState(0) == 0) and (tree.topLevelItem(2).checkState(0) == 0)))

            return res

        tree.itemClicked.connect(lambda: showTable.setDisabled(False) if checkDates() else showTable.setDisabled(True))

        vbox = QVBoxLayout(tab1)
        vbox.addWidget(tree)
        vbox.addWidget(shopList)
        vbox.addWidget(showTable)
        vbox.addStretch(1)
        vbox.addWidget(back)

        tabBar.show()

    def showTableCardWindow(self, shopName, years, months, weeks):
        shopCode = self.controller.getShopCode(shopName)

        print(shopName)
        print([year.text(0) for year in years])
        print([month.text(0) for month in months])
        print([week.text(0) for week in weeks])

        tabBar = QTabWidget(self)
        tabBar.resize(self.width, self.height)
        tab1 = self.currentShopTab
        tab2 = QWidget()
        tabBar.addTab(tab1, self.shopsTabName)
        tabBar.addTab(tab2, self.cardsTabName)

        self.currentCardTab = tab2

        tabBar.setCurrentIndex(1)

        back = QPushButton(self.backButtonName)
        back.setFixedWidth(50)
        back.pressed.connect(lambda: self.shopChooseCardWindow())

        if len(months) != 0 and len(weeks) == 0:
            dates = [month.text(0) for month in months]
            stats = self.controller.getCardStats(shopCode, years=[years[0].text(0)], months=dates)
            time = self.monthStr
        elif len(months) == 0 and len(weeks) != 0:
            dates = [week.text(0) for week in weeks]
            stats = self.controller.getCardStats(shopCode, years=[years[0].text(0)], weeks=dates)
            time = self.weekStr
        else:
            dates = [year.text(0) for year in years]
            stats = self.controller.getCardStats(shopCode, years=dates)
            time = self.yearStr

        table = QTableWidget()
        table.setColumnCount(len(dates))

        # 1 - Количество чеков по картам, 2 - их % от общего числа чеков, 3 - totalSum по картам, 4 - % от общего totalSum
        table.setRowCount(4)

        table.setHorizontalHeaderLabels(dates)
        table.setVerticalHeaderLabels(self.cardStatsLabels)
        header = table.horizontalHeader()
        header.setFrameStyle(QFrame.Box | QFrame.Plain)
        header.setLineWidth(1)
        table.setHorizontalHeader(header)

        for i in range(len(dates)):
            for j in range(4):
                table.setItem(j, i, QTableWidgetItem(str(stats[i][j])))

        table.resizeColumnsToContents()
        table.resizeRowsToContents()
        table.horizontalHeader().setStretchLastSection(True)

        toExcel = QPushButton(self.toExcelButtonName)
        toExcel.pressed.connect(lambda: self.toExcelCard(stats, dates, self.cardStatsLabels, time))

        diagram = QPushButton(self.viewDiagramButtonName)
        diagram.pressed.connect(lambda: self.viewCardDiagram(stats, dates, self.cardStatsLabels, shopName, years, months, weeks))

        vbox = QVBoxLayout(tab2)
        vbox.addWidget(QLabel(shopName))
        vbox.addWidget(table)
        vbox.addWidget(diagram)
        vbox.addWidget(toExcel)
        vbox.addStretch()
        vbox.addWidget(back)

        tabBar.show()

    def showTableShopWindow(self, shopName, years, months, weeks, isCommon):
        if not isCommon:
            shopCode = self.controller.getShopCode(shopName)
        elif shopName != self.controller.wholeCompanyStr:
            city = self.controller.getCityCode(shopName)
        else:
            city = None

        tabBar = QTabWidget(self)
        tabBar.resize(self.width, self.height)
        tab1 = QWidget()
        tab2 = self.currentCardTab
        tabBar.addTab(tab1, self.shopsTabName)
        tabBar.addTab(tab2, self.cardsTabName)

        self.currentShopTab = tab1

        back = QPushButton(self.backButtonName)
        back.setFixedWidth(50)
        back.pressed.connect(lambda: self.shopChooseShopWindow(isCommon))

        if not isCommon:
            if len(months) != 0 and len(weeks) == 0:
                dates = [month.text(0) for month in months]
                stats = self.controller.getShopStats(shopCode, years=[years[0].text(0)], months=dates)
                time = self.monthStr
            elif len(months) == 0 and len(weeks) != 0:
                dates = [week.text(0) for week in weeks]
                stats = self.controller.getShopStats(shopCode, years=[years[0].text(0)], weeks=dates)
                time = self.weekStr
            else:
                dates = [year.text(0) for year in years]
                stats = self.controller.getShopStats(shopCode, years=dates)
                time = self.yearStr
        else:
            if len(months) != 0 and len(weeks) == 0:
                dates = [month.text(0) for month in months]
                stats = self.controller.getCommonShopStats(years=[years[0].text(0)], months=dates, city=city)
                time = self.monthStr
            elif len(months) == 0 and len(weeks) != 0:
                dates = [week.text(0) for week in weeks]
                stats = self.controller.getCommonShopStats(years=[years[0].text(0)], weeks=dates, city=city)
                time = self.weekStr
            else:
                dates = [year.text(0) for year in years]
                stats = self.controller.getCommonShopStats(years=dates, city=city)
                time = self.yearStr

        table = QTableWidget()
        table.setColumnCount(10)
        table.setRowCount(len(dates))

        table.setHorizontalHeaderLabels(self.shopStatsLabels)
        table.setVerticalHeaderLabels(dates)
        header = table.horizontalHeader()
        header.setFrameStyle(QFrame.Box | QFrame.Plain)
        header.setLineWidth(1)
        table.setHorizontalHeader(header)

        for i in range(len(dates)):
            for j in range(10):
                table.setItem(i, j, QTableWidgetItem(str(stats[i][j])))

        table.resizeColumnsToContents()
        table.resizeRowsToContents()
        table.horizontalHeader().setStretchLastSection(True)

        diagram = QPushButton(self.viewDiagramButtonName)
        diagram.pressed.connect(
            lambda: self.viewShopDiagram(stats, dates, self.shopStatsLabels, shopName, years, months, weeks, isCommon))

        if isCommon:
            if len(months) != 0 and len(weeks) == 0:
                manFreqPairsList, manItemList, womanFreqPairsList, womanItemList = \
                    self.controller.getCommonShopFreqStats(years=[years[0].text(0)], months=dates, city=city)
            elif len(months) == 0 and len(weeks) != 0:
                #dates = [month.text(0) for month in months]
                manFreqPairsList, manItemList, womanFreqPairsList, womanItemList = \
                    self.controller.getCommonShopFreqStats(years=[years[0].text(0)], weeks=dates, city=city)
            else:
                #dates = [month.text(0) for month in months]
                manFreqPairsList, manItemList, womanFreqPairsList, womanItemList = \
                    self.controller.getCommonShopFreqStats(years=dates, city=city)
        else:
            if len(months) != 0 and len(weeks) == 0:
                #dates = [month.text(0) for month in months]
                manFreqPairsList, manItemList, womanFreqPairsList, womanItemList = \
                    self.controller.getShopFreqStats(shopCode, years=[years[0].text(0)], months=dates)
            elif len(months) == 0 and len(weeks) != 0:
                #dates = [month.text(0) for month in months]
                manFreqPairsList, manItemList, womanFreqPairsList, womanItemList = \
                    self.controller.getShopFreqStats(shopCode, years=[years[0].text(0)], weeks=dates)
            else:
                #dates = [month.text(0) for month in months]
                manFreqPairsList, manItemList, womanFreqPairsList, womanItemList = \
                    self.controller.getShopFreqStats(shopCode, years=dates)

        toExcel = QPushButton(self.toExcelButtonName)
        toExcel.pressed.connect(lambda: self.toExcelShop(stats, dates, self.shopStatsLabels, time,
                                                         manFreqPairsList, manItemList, womanFreqPairsList, womanItemList))

        manFreqPairsTable = QTableWidget()
        manFreqPairsTable.setRowCount(len(manFreqPairsList))
        manFreqPairsTable.setColumnCount(3)
        manFreqPairsTable.setHorizontalHeaderLabels(self.itemPairsLabels)
        for i in range(len(manFreqPairsList)):
            for j in range(3):
                manFreqPairsTable.setItem(i, j, QTableWidgetItem(str(manFreqPairsList[i][j])))

        manFreqPairsTable.horizontalHeader().setStretchLastSection(True)
        manFreqPairsTable.resizeColumnsToContents()
        manFreqPairsTable.resizeRowsToContents()

        manItemTable = QTableWidget()
        manItemTable.setRowCount(len(manItemList))
        manItemTable.setColumnCount(2)
        manItemTable.setHorizontalHeaderLabels(self.itemLabels)
        for i in range(len(manItemList)):
            for j in range(2):
                manItemTable.setItem(i, j, QTableWidgetItem(str(manItemList[i][j])))

        manItemTable.horizontalHeader().setStretchLastSection(True)
        manItemTable.resizeColumnsToContents()
        manItemTable.resizeRowsToContents()

        womanItemTable = QTableWidget()
        womanItemTable.setRowCount(len(womanItemList))
        womanItemTable.setColumnCount(2)
        womanItemTable.setHorizontalHeaderLabels(self.itemLabels)
        for i in range(len(womanItemList)):
            for j in range(2):
                womanItemTable.setItem(i, j, QTableWidgetItem(str(womanItemList[i][j])))

        womanItemTable.horizontalHeader().setStretchLastSection(True)
        womanItemTable.resizeRowsToContents()
        womanItemTable.resizeColumnsToContents()

        womanFreqPairsTable = QTableWidget()
        womanFreqPairsTable.setRowCount(len(womanFreqPairsList))
        womanFreqPairsTable.setColumnCount(3)
        womanFreqPairsTable.setHorizontalHeaderLabels(self.itemPairsLabels)
        for i in range(len(womanFreqPairsList)):
            for j in range(3):
                womanFreqPairsTable.setItem(i, j, QTableWidgetItem(str(womanFreqPairsList[i][j])))

        womanFreqPairsTable.horizontalHeader().setStretchLastSection(True)
        womanFreqPairsTable.resizeColumnsToContents()
        womanFreqPairsTable.resizeRowsToContents()

        manPairsButton = QPushButton(self.viewDiagramButtonName)
        manPairsButton.pressed.connect(lambda: self.viewCountFreqDiagram(
            [i[2] for i in manFreqPairsList],
            ["(" + str(i[0]) + "," + str(i[1]) + ")" for i in manFreqPairsList],
            shopName, years, months, weeks, isCommon, self.frequencyStr
        ))
        manPairs = QVBoxLayout()
        manPairs.addWidget(manFreqPairsTable)
        manPairs.addWidget(manPairsButton)
        if len(manFreqPairsList) == 0:
            manPairsButton.setDisabled(True)

        manItemsButton = QPushButton(self.viewDiagramButtonName)
        manItemsButton.pressed.connect(lambda: self.viewCountFreqDiagram(
            [i[1] for i in manItemList],
            [str(i[0]) for i in manItemList],
            shopName, years, months, weeks, isCommon, self.countStr
        ))
        manItems = QVBoxLayout()
        manItems.addWidget(manItemTable)
        manItems.addWidget(manItemsButton)
        if len(manItemList) == 0:
            manItemsButton.setDisabled(True)

        womanPairsButton = QPushButton(self.viewDiagramButtonName)
        womanPairsButton.pressed.connect(lambda: self.viewCountFreqDiagram(
            [i[2] for i in womanFreqPairsList],
            ["(" + str(i[0]) + "," + str(i[1]) + ")" for i in womanFreqPairsList],
            shopName, years, months, weeks, isCommon, self.frequencyStr
        ))
        womanPairs = QVBoxLayout()
        womanPairs.addWidget(womanFreqPairsTable)
        womanPairs.addWidget(womanPairsButton)
        if len(womanFreqPairsList) == 0:
            womanPairsButton.setDisabled(True)

        womanItemsButton = QPushButton(self.viewDiagramButtonName)
        womanItemsButton.pressed.connect(lambda: self.viewCountFreqDiagram(
            [i[1] for i in womanItemList],
            [str(i[0]) for i in womanItemList],
            shopName, years, months, weeks, isCommon, self.countStr
        ))
        womanItems = QVBoxLayout()
        womanItems.addWidget(womanItemTable)
        womanItems.addWidget(womanItemsButton)
        if len(womanItemList) == 0:
            womanItemsButton.setDisabled(True)

        itemPairs = QHBoxLayout()
        itemPairs.addLayout(manPairs)
        itemPairs.addLayout(manItems)
        itemPairs.addLayout(womanPairs)
        itemPairs.addLayout(womanItems)

        manWomanLabel = QHBoxLayout()
        manWomanLabel.addWidget(QLabel(self.manLabel))
        manWomanLabel.addWidget(QLabel(self.womanLabel))

        vbox = QVBoxLayout(tab1)
        vbox.addWidget(QLabel(shopName))
        vbox.addWidget(table)
        vbox.addWidget(diagram)
        vbox.addLayout(manWomanLabel)
        vbox.addLayout(itemPairs)
        vbox.addWidget(toExcel)
        vbox.addStretch()
        vbox.addWidget(back)

        tabBar.show()

    def viewCountFreqDiagram(self, stats, elements, shopName, years, months, days, isCommon, yLabel):
        stats = [int(i) for i in stats]

        global currentShopPlot

        def drawPlot():
            global currentShopPlot

            width = 1 / (len(elements))
            index = np.arange(len(elements))
            xs = []
            for i in range(len(elements)):
                xs.append(stats[i])

            bar(index, xs, width=width, zorder=2)

            xticks(index, elements, rotation=30, ha="right")
            ylabel(yLabel)
            tight_layout()
            grid(axis='y')

            buf = io.BytesIO()

            gcf().set_size_inches(10.0, 6.4)

            savefig(buf, format='png', dpi=100)
            buf.seek(0)

            im = Image.open(buf)

            imshow(im)
            buf.close()

            currentShopPlot = im

            pic1 = QLabel()
            pic1.setGeometry(0, 0, 640, 480)

            pic1.setPixmap(QPixmap.fromImage(ImageQt(im)))

            clf()
            return pic1

        tabBar = QTabWidget(self)
        tabBar.resize(self.width, self.height)
        tab1 = QWidget()
        tab2 = self.currentCardTab
        tabBar.addTab(tab1, self.shopsTabName)
        tabBar.addTab(tab2, self.cardsTabName)

        self.currentShopTab = tab1

        pic = drawPlot()

        toFile = QPushButton(self.toPngButtonName)
        toFile.pressed.connect(lambda: self.toFile(currentShopPlot))

        back = QPushButton(self.backButtonName)
        back.setFixedWidth(50)
        back.pressed.connect(lambda: self.showTableShopWindow(shopName, years, months, days, isCommon))

        vbox = QVBoxLayout(tab1)
        vbox.addWidget(pic)
        vbox.addWidget(toFile)
        vbox.addStretch()
        vbox.addWidget(back)

        tabBar.show()

    def viewCardDiagram(self, stats, dates, labels, shopName, years, months, days, indexOfParameter=0):
        hbox = QHBoxLayout()
        global currentCardPlot

        def drawPlot(indexOfParameter):
            global currentCardPlot

            width = 1 / (len(dates))
            index = np.arange(len(dates))
            xs = []
            for i in range(len(dates)):
                xs.append(stats[i][indexOfParameter])

            if indexOfParameter == 0 or indexOfParameter == 2:
                bar(index, xs, width=width, zorder=2)

            elif indexOfParameter == 1 or indexOfParameter == 3:
                bar(index, [1 for i in range(len(dates))], width=width, zorder=2)
                bar(index, xs, width=width, zorder=2)

                legend(["All", "Cards"], loc="upper left")

            xticks(index, dates, rotation=30, ha="right")
            ylabel(labels[indexOfParameter])
            tight_layout()
            grid(axis='y')

            buf = io.BytesIO()

            gcf().set_size_inches(8.0, 6.4)

            savefig(buf, format='png', dpi=100)
            buf.seek(0)

            im = Image.open(buf)

            imshow(im)
            buf.close()

            currentCardPlot = im

            pic1 = QLabel()
            pic1.setGeometry(0, 0, 640, 480)

            pic1.setPixmap(QPixmap.fromImage(ImageQt(im)))

            clf()
            return pic1

        tabBar = QTabWidget(self)
        tabBar.resize(self.width, self.height)
        tab1 = self.currentShopTab
        tab2 = QWidget()
        tabBar.addTab(tab1, self.shopsTabName)
        tabBar.addTab(tab2, self.cardsTabName)

        tabBar.setCurrentIndex(1)

        self.currentCardTab = tab2

        pic = drawPlot(indexOfParameter)

        chooseLabel = QLabel(self.chooseIndexStr)
        chooseBox = QComboBox()
        chooseBox.addItems(labels)
        chooseBox.setFixedWidth(350)

        hbox.addWidget(pic)
        hbox.addStretch()
        hbox.addWidget(chooseLabel)
        hbox.addWidget(chooseBox)

        chooseBox.currentIndexChanged.connect(
            lambda: {hbox.insertWidget(0, drawPlot(chooseBox.currentIndex())), hbox.takeAt(1)})

        back = QPushButton(self.backButtonName)
        back.setFixedWidth(50)
        back.pressed.connect(lambda: self.showTableCardWindow(shopName, years, months, days))

        toFile = QPushButton(self.toPngButtonName)
        toFile.pressed.connect(lambda: self.toFile(currentCardPlot))

        vbox = QVBoxLayout(tab2)
        vbox.addLayout(hbox)
        vbox.addWidget(toFile)
        vbox.addStretch()
        vbox.addWidget(back)

        tabBar.show()

    def viewShopDiagram(self, stats, dates, labels, shopName, years, months, days, isCommon, numberOfKPI=0):
        hbox = QHBoxLayout()
        global currentShopPlot

        def drawPlot(numberOfKPI):
            global currentShopPlot

            width = 1 / (len(dates))
            index = np.arange(len(dates))
            xs = []
            for i in range(len(dates)):
                xs.append(stats[i][numberOfKPI])

            bar(index, xs, width=width, zorder=2)

            xticks(index, dates, rotation=30, ha="right")
            ylabel(labels[numberOfKPI])
            tight_layout()
            grid(axis='y')

            buf = io.BytesIO()

            gcf().set_size_inches(9.0, 6.0)

            savefig(buf, format='png', dpi=100)
            buf.seek(0)

            im = Image.open(buf)

            imshow(im)
            buf.close()

            currentShopPlot = im

            pic1 = QLabel()
            pic1.setGeometry(0, 0, 640, 480)


            pic1.setPixmap(QPixmap.fromImage(ImageQt(im)))

            clf()
            return pic1

        tabBar = QTabWidget(self)
        tabBar.resize(self.width, self.height)
        tab1 = QWidget()
        tab2 = self.currentCardTab
        tabBar.addTab(tab1, self.shopsTabName)
        tabBar.addTab(tab2, self.cardsTabName)

        self.currentShopTab = tab1

        pic = drawPlot(numberOfKPI)

        chooseLabel = QLabel(self.chooseIndexStr)
        chooseBox = QComboBox()
        chooseBox.addItems(labels)

        hbox.addWidget(pic)
        hbox.addStretch()
        hbox.addWidget(chooseLabel)
        hbox.addWidget(chooseBox)

        chooseBox.currentIndexChanged.connect(lambda: {hbox.insertWidget(0, drawPlot(chooseBox.currentIndex())), hbox.takeAt(1)})

        back = QPushButton(self.backButtonName)
        back.setFixedWidth(50)
        back.pressed.connect(lambda: self.showTableShopWindow(shopName, years, months, days, isCommon))

        toFile = QPushButton(self.toPngButtonName)
        toFile.pressed.connect(lambda: self.toFile(currentShopPlot))

        vbox = QVBoxLayout(tab1)
        vbox.addLayout(hbox)
        vbox.addWidget(toFile)
        vbox.addStretch()
        vbox.addWidget(back)

        tabBar.show()

    def toFile(self, plot):
        path = QFileDialog().getExistingDirectory(self, self.chooseDirectoryStr)

        if path:

            text, ok = QInputDialog.getText(self, self.chooseNameStr, self.enterPngNameStr)

            while text == "" and ok:
                error = QMessageBox()
                error.setIcon(QMessageBox.Critical)
                error.setText(self.fileNameNotChosenMsg)
                error.setStandardButtons(QMessageBox.Ok)
                error.exec_()

                text, ok = QInputDialog.getText(self, self.chooseNameStr, self.enterPngNameStr)

            if ok:
                plot.save(path + '/' + text + ".png")

    def toExcelShop(self, stats, dates, labels, timeName, manFreqPairsList, manItemList, womanFreqPairsList, womanItemList):
        path = QFileDialog().getExistingDirectory(self, self.chooseDirectoryStr)

        if path:

            text, ok = QInputDialog.getText(self, self.chooseNameStr, self.enterExcelNameStr)

            while text == "" and ok:
                error = QMessageBox()
                error.setIcon(QMessageBox.Critical)
                error.setText(self.fileNameNotChosenMsg)
                error.setStandardButtons(QMessageBox.Ok)
                error.exec_()

                text, ok = QInputDialog.getText(self, self.chooseNameStr, self.enterExcelNameStr)

            if ok:
                fileStats = openpyxl.Workbook()

                fileStats.create_sheet('Stats', 0)
                fileStats.create_sheet('Man Pairs', 1)
                fileStats.create_sheet('Man Items', 2)
                fileStats.create_sheet('Woman Pairs', 3)
                fileStats.create_sheet('Woman Items', 4)

                ws = fileStats['Stats']

                ws.cell(row=1, column=1).value = timeName

                for i in range(len(dates)):
                    ws.cell(row=i + 2, column=1).value = int(dates[i])

                for i in range(len(labels)):
                    ws.cell(row=1, column=i + 2).value = labels[i]

                for i in range(len(stats)):
                    for j in range(len(stats[i])):
                        ws.cell(row=i + 2, column=j + 2).value = stats[i][j]

                ws = fileStats['Man Pairs']

                ws.cell(row=1, column=1).value = "First Item"
                ws.cell(row=1, column=2).value = "Second Item"
                ws.cell(row=1, column=3).value = "Frequency"

                for i in range(len(manFreqPairsList)):
                    for j in range(len(manFreqPairsList[i])):
                        ws.cell(row=i + 2, column=j + 1).value = str(manFreqPairsList[i][j])

                ws = fileStats['Man Items']

                ws.cell(row=1, column=1).value = "Item"
                ws.cell(row=1, column=2).value = "Count"

                for i in range(len(manItemList)):
                    for j in range(len(manItemList[i])):
                        ws.cell(row=i + 2, column=j + 1).value = str(manItemList[i][j])

                ws = fileStats['Woman Pairs']

                ws.cell(row=1, column=1).value = "First Item"
                ws.cell(row=1, column=2).value = "Second Item"
                ws.cell(row=1, column=3).value = "Frequency"

                for i in range(len(womanFreqPairsList)):
                    for j in range(len(womanFreqPairsList[i])):
                        ws.cell(row=i + 2, column=j + 1).value = str(womanFreqPairsList[i][j])

                ws = fileStats['Woman Items']

                ws.cell(row=1, column=1).value = "Item"
                ws.cell(row=1, column=2).value = "Count"

                for i in range(len(womanItemList)):
                    for j in range(len(womanItemList[i])):
                        ws.cell(row=i + 2, column=j + 1).value = str(womanItemList[i][j])

                fileStats.save(path + '/' + text + ".xlsx")

    def toExcelCard(self, stats, dates, labels, timeName):
        path = QFileDialog().getExistingDirectory(self, self.chooseDirectoryStr)
        if path:

            text, ok = QInputDialog.getText(self, self.chooseNameStr, self.enterExcelNameStr)

            while text == "" and ok:
                error = QMessageBox()
                error.setIcon(QMessageBox.Critical)
                error.setText(self.fileNameNotChosenMsg)
                error.setStandardButtons(QMessageBox.Ok)
                error.exec_()

                text, ok = QInputDialog.getText(self, self.chooseNameStr, self.enterExcelNameStr)

            if ok:
                fileStats = openpyxl.Workbook()
                fileStats.create_sheet('Stats', 0)
                ws = fileStats['Stats']

                for i in range(len(dates)):
                    ws.cell(row=i + 2, column=1).value = int(dates[i])

                for i in range(len(labels)):
                    ws.cell(row=1, column=i + 2).value = labels[i]

                for i in range(len(stats)):
                    for j in range(len(stats[i])):
                        ws.cell(row=i + 2, column=j + 2).value = stats[i][j]

                ws.cell(row=1, column=1).value = timeName

                fileStats.save(path + '/' + text + ".xlsx")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = Window()
    sys.exit(app.exec_())
